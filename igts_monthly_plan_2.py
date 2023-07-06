import openpyxl
from datetime import datetime

CATEGORIES = [
    'Financial goals',
    'Tools',
    'Deadlines',
    'Plan',
    'Measuring progress'
]

def prompt_user():
    answers = []
    for category in CATEGORIES:
        if category == 'Deadlines':
            answer = input(f"3. {category}: ")
            while not answer.isnumeric():
                answer = input(f"Please enter a numeric deadline: ")
            answers.append(int(answer))
        elif category == 'Financial goals':
            goal1 = input("Enter your first financial goal for this month: ")
            goal1_deadline = input(f"When would you like to achieve {goal1}? ")
            goal1_action = input(f"What action do you need to take to achieve {goal1}? ")
            goal1_contact = input(f"Who is a good person to call before taking action on {goal1}? ")
            goal2 = input("Enter your second financial goal for this month: ")
            goal2_deadline = input(f"When would you like to achieve {goal2}? ")
            goal2_action = input(f"What action do you need to take to achieve {goal2}? ")
            goal2_contact = input(f"Who is a good person to call before taking action on {goal2}? ")
            goal3 = input("Enter your third financial goal for this month: ")
            goal3_deadline = input(f"When would you like to achieve {goal3}? ")
            goal3_action = input(f"What action do you need to take to achieve {goal3}? ")
            goal3_contact = input(f"Who is a good person to call before taking action on {goal3}? ")
            answers.append(
                f"{goal1}, {goal2}, {goal3}"
            )
            answers.append(
                [
                    {'goal': goal1, 'deadline': goal1_deadline, 'action': goal1_action, 'contact': goal1_contact},
                    {'goal': goal2, 'deadline': goal2_deadline, 'action': goal2_action, 'contact': goal2_contact},
                    {'goal': goal3, 'deadline': goal3_deadline, 'action': goal3_action, 'contact': goal3_contact},
                ]
            )
        elif category == 'Measuring progress':
            progress = input("Enter your progress for this month (out of 100): ")
            while not progress.isnumeric() or int(progress) < 0 or int(progress) > 100:
                progress = input("Please enter a valid progress percentage (0-100): ")
            answers.append(int(progress))
        else:
            answer = input(f"{category}: ")
            answers.append(answer)

    # Collect friend's name and meal times
    friend_name = input("Enter your friend's name: ")
    answers.append(friend_name)
    for category in ['Coffee', 'Breakfast', 'Lunch', 'Afternoon Snack', 'Dinner']:
        times = []
        for i in range(1, 6):
            time = input(f"Enter the best time for {category} for person {i}: ")
            times.append(time)
        answers.append(times)

    # Collect schedule
    schedule = {}
    month = input("Enter the month (1-12): ")
    year = input("Enter the year: ")
    for i in range(1, 32):
        try:
            date = datetime.strptime(f"{i}/{month}/{year}", '%d/%m/%Y').strftime('%d/%m/%Y')
        except ValueError:
            break
        schedule[date] = {}
        for j, category in enumerate(['Coffee', 'Breakfast', 'Lunch', 'Afternoon Snack', 'Dinner']):
            times = []
            for k, person in enumerate(['Person 1', 'Person 2', 'Person 3', 'Person 4', 'Person 5']):
                time = input(f"Enter the best time for {category} for {person} on {date}: ")
                times.append(time)
            schedule[date][category] = times

        answers.append(schedule)

    return answers


def save_to_excel(answers):
    workbook = openpyxl.Workbook()

    # Save financial goals and plan
    sheet1 = workbook.active
    sheet1.title = "Goals and Plan"
    sheet1.cell(row=1, column=1, value="Category")
    sheet1.cell(row=1, column=2, value="Answer")
    for i, category in enumerate(CATEGORIES):
        sheet1.cell(row=i+2, column=1, value=category)
        if category == 'Plan':
            sheet1.cell(row=i+2,column=2, value='\n'.join(answers[i]))
        elif category == 'Financial goals':
            goals = answers[i]
            for j, goal in enumerate(goals):
                sheet1.cell(row=i+j+2, column=2, value=goal)
            sheet1.cell(row=i+2, column=3, value='Deadline')
            sheet1.cell(row=i+2, column=4, value='Action')
            sheet1.cell(row=i+2, column=5, value='Contact')
            for j, goal in enumerate(answers[i+1]):
                sheet1.cell(row=i+j+3, column=3, value=goal['deadline'])
                sheet1.cell(row=i+j+3, column=4, value=goal['action'])
                sheet1.cell(row=i+j+3, column=5, value=goal['contact'])
        else:
            sheet1.cell(row=i+2, column=2, value=answers[i])

    # Save progress
    sheet2 = workbook.create_sheet(title="Progress")
    sheet2.cell(row=1, column=1, value="Category")
    sheet2.cell(row=1, column=2, value="Answer")
    for i, category in enumerate(CATEGORIES):
        sheet2.cell(row=i+2, column=1, value=category)
        if category == 'Measuring progress':
            sheet2.cell(row=i+2, column=2, value=f"{answers[i]}%")
        else:
            sheet2.cell(row=i+2, column=2, value=answers[i])

    # Save schedule
    schedule = answers[-1]
    dates = list(schedule.keys())
    dates.sort(key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
    for date in dates:
        sheet3 = workbook.create_sheet(title=date)
        sheet3.cell(row=1, column=1, value="Category")
        sheet3.cell(row=1, column=2, value="Person 1")
        sheet3.cell(row=1, column=3, value="Person 2")
        sheet3.cell(row=1, column=4, value="Person 3")
        sheet3.cell(row=1, column=5, value="Person 4")
        sheet3.cell(row=1, column=6, value="Person 5")
        for i, category in enumerate(['Coffee', 'Breakfast', 'Lunch', 'Afternoon Snack', 'Dinner']):
            sheet3.cell(row=i+2, column=1, value=category)
            times = schedule[date][category]
            for j, time in enumerate(times):
                sheet3.cell(row=i+2, column=j+2, value=time)

    workbook.save(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_planner.xlsx")


if __name__ == '__main__':
    answers = prompt_user()
    save_to_excel(answers)
    print("Data saved to Excel file.")