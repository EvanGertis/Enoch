import openpyxl
from datetime import datetime

CATEGORIES = [
    'Financial and health goals',
    'Tools and resources',
    'Deadlines',
    'Plan of action',
    'Measuring progress',
    'Schedule advice'
]

def prompt_user():
    answers = []
    for category in CATEGORIES:
        if category == 'Deadlines':
            answer = input(f"3. {category}: ")
            while not answer.isnumeric():
                answer = input(f"Please enter a numeric deadline: ")
            answers.append(int(answer))
        elif category == 'Financial and health goals':
            financial_goal1 = input("Enter your first financial goal for this month: ")
            financial_goal1_deadline = input(f"When would you like to achieve {financial_goal1}? ")
            financial_goal1_action = input(f"What action do you need to take to achieve {financial_goal1}? ")
            financial_goal1_contact = input(f"Who is a good person to call before taking action on {financial_goal1}? ")
            financial_goal2 = input("Enter your second financial goal for this month: ")
            financial_goal2_deadline = input(f"When would you like to achieve {financial_goal2}? ")
            financial_goal2_action = input(f"What action do you need to take to achieve {financial_goal2}? ")
            financial_goal2_contact = input(f"Who is a good person to call before taking action on {financial_goal2}? ")
            financial_goal3 = input("Enter your third financial goal for this month: ")
            financial_goal3_deadline = input(f"When would you like to achieve {financial_goal3}? ")
            financial_goal3_action = input(f"What action do you need to take to achieve {financial_goal3}? ")
            financial_goal3_contact = input(f"Who is a good person to call before taking action on {financial_goal3}? ")
            health_goal1 = input("Enter your first health goal for this month: ")
            health_goal1_deadline = input(f"When would you like to achieve {health_goal1}? ")
            health_goal1_action = input(f"What action do you need to take to achieve {health_goal1}? ")
            health_goal1_contact = input(f"Who is a good person to call before taking action on {health_goal1}? ")
            health_goal2 = input("Enter your second health goal for this month: ")
            health_goal2_deadline = input(f"When would you like to achieve {health_goal2}? ")
            health_goal2_action = input(f"What action do you need to take to achieve {health_goal2}? ")
            health_goal2_contact = input(f"Who is a good person to call before taking action on {health_goal2}? ")
            health_goal3 = input("Enter your third health goal for this month: ")
            health_goal3_deadline = input(f"When would you like to achieve {health_goal3}? ")
            health_goal3_action = input(f"What action do you need to take to achieve {health_goal3}? ")
            health_goal3_contact = input(f"Who is a good person to call before taking action on {health_goal3}? ")
            answers.append(
                [
                    {'goal': financial_goal1, 'deadline': financial_goal1_deadline, 'action': financial_goal1_action, 'contact': financial_goal1_contact},
                    {'goal': financial_goal2, 'deadline': financial_goal2_deadline, 'action': financial_goal2_action, 'contact': financial_goal2_contact},
                    {'goal': financial_goal3, 'deadline': financial_goal3_deadline, 'action': financial_goal3_action, 'contact': financial_goal3_contact},
                    {'goal': health_goal1, 'deadline': health_goal1_deadline, 'action': health_goal1_action, 'contact': health_goal1_contact},
                    {'goal': health_goal2, 'deadline': health_goal2_deadline, 'action': health_goal2_action, 'contact': health_goal2_contact},
                    {'goal': health_goal3, 'deadline': health_goal3_deadline, 'action': health_goal3_action, 'contact': health_goal3_contact}
                ]
            )
        elif category == 'Tools and resources':
            financial_tools = input("Enter the financial management tools and investment services available to you: ")
            financial_tools_support = input(f"Who can you reach out to for support with {financial_tools}? ")
            health_tools = input("Enter the health tracking tools available to you: ")
            health_tools_support = input(f"Who can you reach out to for support with {health_tools}? ")
            answers.append(
                {'financial_tools': financial_tools, 'financial_tools_support': financial_tools_support, 'health_tools': health_tools, 'health_tools_support': health_tools_support}
            )
        elif category == 'Deadlines':
            answer = input(f"3. {category}: ")
            while not answer.isnumeric():
               answer = input(f"Please enter a numeric deadline: ")
            answers.append(int(answer))
        elif category == 'Plan of action':
            financial_plan = input("What is your plan of action for achieving your financial goals this month? ")
            health_plan = input("What is your plan of action for achieving your health goals this month? ")
            answers.append(
                {'financial_plan': financial_plan, 'health_plan': health_plan}
            )
        elif category == 'Measuring progress':
            financial_progress = input("How will you measure your progress towards your financial goals? ")
            health_progress = input("How will you measure your progress towards your health goals? ")
            answers.append(
                {'financial_progress': financial_progress, 'health_progress': health_progress}
            )
        elif category == 'Schedule advice':
            schedule_advice = input("What advice do you have for scheduling your day to achieve your goals? ")
            answers.append(schedule_advice)

    return answers

def save_to_workbook(answers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Goal Setting Worksheet"
    ws['A1'] = "Goal Setting Worksheet"
    ws['A2'] = f"Created on {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}"
    row = 4
    for idx, category in enumerate(CATEGORIES):
        ws.cell(row=row, column=1, value=category)
        if category == 'Financial and health goals':
            for goal in answers[idx]:
                ws.cell(row=row+1, column=1, value=goal['goal'])
                ws.cell(row=row+1, column=2, value=goal['deadline'])
                ws.cell(row=row+1, column=3, value=goal['action'])
                ws.cell(row=row+1, column=4, value=goal['contact'])
                row += 1
        elif category == 'Tools and resources':
            ws.cell(row=row+1, column=1, value='Financial Management Tools and Investment Services')
            ws.cell(row=row+1, column=2, value=answers[idx]['financial_tools'])
            ws.cell(row=row+1, column=3, value=answers[idx]['financial_tools_support'])
            ws.cell(row=row+2, column=1, value='Health Tracking Tools')
            ws.cell(row=row+2, column=2, value=answers[idx]['health_tools'])
            ws.cell(row=row+2, column=3, value=answers[idx]['health_tools_support'])
            row += 3
        elif category == 'Plan of action':
            ws.cell(row=row+1, column=1, value='Financial Plan of Action')
            ws.cell(row=row+1, column=2, value=answers[idx]['financial_plan'])
            ws.cell(row=row+2, column=1, value='Health Plan of Action')
            ws.cell(row=row+2, column=2, value=answers[idx]['health_plan'])
            row += 3
        elif category == 'Measuring progress':
            ws.cell(row=row+1, column=1, value='Measuring Progress Towards Financial Goals')
            ws.cell(row=row+1, column=2, value=answers[idx]['financial_progress'])
            ws.cell(row=row+2, column=1, value='Measuring Progress Towards Health Goals')
            ws.cell(row=row+2, column=2, value=answers[idx]['health_progress'])
            row += 3
        else:
            ws.cell(row=row+1, column=1, value=answers[idx])
            row += 2

    filename = f"Goal Setting Worksheet {datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.xlsx"
    wb.save(filename)
    print(f"Your goal setting worksheet has been saved to {filename}")

if __name__ == '__main__':
    print("Welcome to the Goal Setting Worksheet!")
    print("Please answer the following questions to help you set and achieve your goals this month:")
    answers = prompt_user()
    save_to_workbook(answers)