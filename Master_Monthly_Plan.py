import openpyxl
from datetime import datetime

CATEGORIES = [
    'Financial and health goals',
    'Tools and resources',
    'Deadlines',
    'Plan of action',
    'Measuring progress',
    'Schedule advice'
]

def prompt_user():
    answers = []
    for category in CATEGORIES:
        if category == 'Deadlines':
            answer = input(f"3. {category}: ")
            while True:
                try:
                    deadline = datetime.strptime(answer, '%m/%d/%Y')
                    break
                except ValueError:
                    answer = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers.append(deadline.date())
        elif category == 'Financial and health goals':
            financial_goal1 = input("Enter your first financial goal for this month: ")
            financial_goal1_deadline = input(f"When would you like to achieve {financial_goal1}? (MM/DD/YYYY) ")
            financial_goal1_action = input(f"What action do you need to take to achieve {financial_goal1}? ")
            financial_goal1_contact = input(f"Who is a good person to call before taking action on {financial_goal1}? ")
            while True:
                try:
                    deadline = datetime.strptime(financial_goal1_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    financial_goal1_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers.append(
                [
                    {'goal': financial_goal1, 'deadline': deadline.date(), 'action': financial_goal1_action, 'contact': financial_goal1_contact},
                ]
            )
            financial_goal2 = input("Enter your second financial goal for this month: ")
            financial_goal2_deadline = input(f"When would you like to achieve {financial_goal2}? (MM/DD/YYYY) ")
            financial_goal2_action = input(f"What action do you need to take to achieve {financial_goal2}? ")
            financial_goal2_contact = input(f"Who is a good person to call before taking action on {financial_goal2}? ")
            while True:
                try:
                    deadline = datetime.strptime(financial_goal2_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    financial_goal2_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers[-1].append(
                {'goal': financial_goal2, 'deadline': deadline.date(), 'action': financial_goal2_action, 'contact': financial_goal2_contact},
            )
            financial_goal3 = input("Enter your third financial goal for this month: ")
            financial_goal3_deadline = input(f"When would you like to achieve {financial_goal3}? (MM/DD/YYYY) ")
            financial_goal3_action = input(f"What action do you need to take to achieve {financial_goal3}? ")
            financial_goal3_contact = input(f"Who is a good person to call before taking action on {financial_goal3}? ")
            while True:
                try:
                    deadline = datetime.strptime(financial_goal3_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    financial_goal3_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers[-1].append(
                {'goal': financial_goal3, 'deadline': deadline.date(), 'action': financial_goal3_action, 'contact': financial_goal3_contact},
            )
            health_goal1 = input("Enter your first health goal for this month: ")
            health_goal1_deadline = input(f"When would you like to achieve {health_goal1}? (MM/DD/YYYY) ")
            health_goal1_action = input(f"What action do you need to take to achieve {health_goal1}? ")
            health_goal1_contact = input(f"Who is a good person to call before taking action on {health_goal1}? ")
            while True:
                try:
                    deadline = datetime.strptime(health_goal1_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    health_goal1_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers[-1].append(
                {'goal': health_goal1, 'deadline': deadline.date(), 'action': health_goal1_action, 'contact': health_goal1_contact},
            )
            health_goal2 = input("Enter your second health goal for this month: ")
            health_goal2_deadline = input(f"When would you like to achieve {health_goal2}? (MM/DD/YYYY) ")
            health_goal2_action = input(f"What action do you need to take to achieve {health_goal2}? ")
            health_goal2_contact = input(f"Who is a good person to call before taking action on {health_goal2}? ")
            while True:
                try:
                    deadline = datetime.strptime(health_goal2_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    health_goal2_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers[-1].append(
                {'goal': health_goal2, 'deadline': deadline.date(), 'action': health_goal2_action, 'contact': health_goal2_contact},
            )
            health_goal3 = input("Enter your third health goal for this month: ")
            health_goal3_deadline = input(f"When would you like to achieve {health_goal3}? (MM/DD/YYYY) ")
            health_goal3_action = input(f"What action do you need to take to achieve {health_goal3}? ")
            health_goal3_contact = input(f"Who is a good person to call before taking action on {health_goal3}? ")
            while True:
                try:
                    deadline = datetime.strptime(health_goal3_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    health_goal3_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers[-1].append(
                {'goal': health_goal3, 'deadline': deadline.date(), 'action': health_goal3_action, 'contact': health_goal3_contact},
            )
        else:
            answer = input(f"3. {category}: ")
            answers.append(answer)
    return answers

def save_to_workbook(user_input):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Monthly Goals'
    sheet['A1'] = 'Monthly Goals'
    sheet['A2'] = 'Category'
    sheet['B2'] = 'Answer'

    row = 3
    for category, answer in zip(CATEGORIES, user_input):
        if category == 'Financial and health goals':
            sheet.cell(row=row, column=1).value = category
            sheet.cell(row=row, column=2).value = None
            sheet.cell(row=row, column=3).value = 'Goal'
            sheet.cell(row=row, column=4).value = 'When to achieve'
            sheet.cell(row=row, column=5).value = 'Action to be taken'
            sheet.cell(row=row, column=6).value = 'Who to call for advice'
            row += 1
            for goal in answer:
                sheet.cell(row=row, column=1).value = None
                sheet.cell(row=row, column=2).value = None
                sheet.cell(row=row, column=3).value = goal['goal']
                sheet.cell(row=row, column=4).value = goal['deadline']
                sheet.cell(row=row, column=5).value = goal['action']
                sheet.cell(row=row, column=6).value = goal['contact']
                row += 1
        else:
            sheet.cell(row=row, column=1).value = category
            sheet.cell(row=row, column=2).value = answer
            row += 1

    wb.save('monthly_goals.xlsx')
    print("Monthly goals saved to monthly_goals.xlsx")

if __name__ == '__main__':
    user_input = prompt_user()
    save_to_workbook(user_input)