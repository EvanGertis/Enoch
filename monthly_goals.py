import openpyxl
from datetime import datetime

CATEGORIES = [
    'Financial and health goals',
    'Tools and resources',
    'Deadlines',
    'Plan of action',
    'Measuring progress',
    'Schedule advice'
]

def prompt_user():
    answers = []
    for category in CATEGORIES:
        if category == 'Deadlines':
            answer = input(f"3. {category}: ")
            while True:
                try:
                    deadline = datetime.strptime(answer, '%m/%d/%Y')
                    break
                except ValueError:
                    answer = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers.append(deadline.date())
        elif category == 'Financial and health goals':
            financial_goal1 = input("Enter your first financial goal for this month: ")
            financial_goal1_deadline = input(f"When would you like to achieve {financial_goal1}? (MM/DD/YYYY) ")
            financial_goal1_action = input(f"What action do you need to take to achieve {financial_goal1}? ")
            financial_goal1_contact = input(f"Who is a good person to call before taking action on {financial_goal1}? ")
            while True:
                try:
                    deadline = datetime.strptime(financial_goal1_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    financial_goal1_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers.append(
                [
                    {'goal': financial_goal1, 'deadline': deadline.date(), 'action': financial_goal1_action, 'contact': financial_goal1_contact, 'progress': []},
                ]
            )
            financial_goal2 = input("Enter your second financial goal for this month: ")
            financial_goal2_deadline = input(f"When would you like to achieve {financial_goal2}? (MM/DD/YYYY) ")
            financial_goal2_action = input(f"What action do you need to take to achieve {financial_goal2}? ")
            financial_goal2_contact = input(f"Who is a good person to call before taking action on {financial_goal2}? ")
            while True:
                try:
                    deadline = datetime.strptime(financial_goal2_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    financial_goal2_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers[-1].append(
                {'goal': financial_goal2, 'deadline': deadline.date(), 'action': financial_goal2_action, 'contact': financial_goal2_contact, 'progress': []},
            )
            financial_goal3 = input("Enter your third financial goal for this month: ")
            financial_goal3_deadline = input(f"When would you like to achieve {financial_goal3}? (MM/DD/YYYY) ")
            financial_goal3_action = input(f"What action do you need to take to achieve {financial_goal3}? ")
            financial_goal3_contact = input(f"Who is a good person to call before taking action on {financial_goal3}? ")
            while True:
                try:
                    deadline = datetime.strptime(financial_goal3_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    financial_goal3_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers[-1].append(
                {'goal': financial_goal3, 'deadline': deadline.date(), 'action': financial_goal3_action, 'contact': financial_goal3_contact, 'progress': []},
            )
            health_goal1 = input("Enter your first health goal for this month: ")
            health_goal1_deadline = input(f"When would you like to achieve {health_goal1}? (MM/DD/YYYY) ")
            health_goal1_action = input(f"What action do you need to take to achieve {health_goal1}? ")
            health_goal1_contact = input(f"Who is a good person to call before taking action on {health_goal1}? ")
            while True:
                try:
                    deadline = datetime.strptime(health_goal1_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    health_goal1_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers[-1].append(
                {'goal': health_goal1, 'deadline': deadline.date(), 'action': health_goal1_action, 'contact': health_goal1_contact, 'progress': []},
            )
            health_goal2 = input("Enter your second health goal for this month: ")
            health_goal2_deadline = input(f"When would you like to achieve {health_goal2}? (MM/DD/YYYY) ")
            health_goal2_action = input(f"What action do you need to take to achieve {health_goal2}? ")
            health_goal2_contact = input(f"Who is a good person to call before taking action on {health_goal2}? ")
            while True:
                try:
                    deadline = datetime.strptime(health_goal2_deadline, '%m/%d/%Y')
                    break
                except ValueError:
                    health_goal2_deadline = input(f"Please enter a deadline in the format MM/DD/YYYY: ")
            answers[-1].append(
                {'goal': health_goal2, 'deadline': deadline.date(), 'action': health_goal2_action, 'contact': health_goal2_contact, 'progress': []},
            )
        elif category == 'Measuring progress':
            for goal in answers[-2]:
                goal_name = goal['goal']
                goal_progress = []
                for i in range(3):
                    progress = input(f"What is your progress on {goal_name} for today? (Day {i+1}) ")
                    goal_progress.append(progress)
                goal['progress'].append({'date': datetime.now().date(), 'progress': goal_progress})
            for goal in answers[-1]:
                goal_name = goal['goal']
                goal_progress = []
                for i in range(3):
                    progress = input(f"What is your progress on {goal_name} for today? (Day {i+1}) ")
                    goal_progress.append(progress)
                goal['progress'].append({'date': datetime.now().date(), 'progress': goal_progress})
        else:
            answer = input(f"3. {category}: ")
            answers.append(answer)
    return answers

def save_answers(answers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Goals"
    ws['A1'] = "Category"
    ws['B1'] = "Answer"
    row = 2
    for category, answer in zip(CATEGORIES, answers):
        if category == 'Financial and health goals':
            for goal in answer:
                ws.cell(row=row, column=1).value = f"{category}: {goal['goal']}"
                ws.cell(row=row, column=2).value = f"Deadline: {goal['deadline']}\nAction: {goal['action']}\nContact: {goal['contact']}\nProgress: {goal['progress']}"
                row += 1
        else:
            ws.cell(row=row, column=1).value = category
            ws.cell(row=row, column=2).value = answer
            row += 1
    wb.save("monthly_goals.xlsx")

if __name__ == '__main__':
    answers = prompt_user()
    save_answers(answers)
    
    print("Your answers have been saved to monthly_goals.xlsx")