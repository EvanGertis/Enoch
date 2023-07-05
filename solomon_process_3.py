import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

CATEGORIES = {
    "Emergency Fund": 0.1,
    "Brokerage Account": 0.05,
    "Vacations": 0.05,
    "Entertainment": 0.08,
    "Restaurants": 0.05,
    "Clothing": 0.038,
    "Family": 0.05,
    "Christmas": 0.02,
    "Gifts": 0.02,
    "Hobbies": 0.1,
    "House Upgrades": 0.05,
    "Cable": 0.02,
    "Internet": 0.02,
    "Subscriptions": 0.02,
    "Lawn": 0.02,
    "Beauty Products": 0.02,
    "Gym Memberships": 0.02,
    "Mortgage": 0.18,
    "House Taxes": 0.05,
    "House Insurance": 0.02,
    "Electricity": 0.18,
    "House gas": 0.02,
    "Water / Garbage": 0.02,
    "House Repairs": 0.05,
    "Other housing costs": 0.05,
    "Groceries": 0.1,
    "Vehicle Gas": 0.02,
    "Grandkids": 0.02,
    "Vehicle Taxes and Insurance": 0.05,
    "Vehicle Replacement": 0.05,
    "Oil Changes": 0.02,
    "AAA": 0.02,
    "Tires": 0.02,
    "Vehicle Repairs / Upkeep": 0.05,
    "Health Insurance Premiums": 0.1,
    "Healthcare Costs": 0.05,
    "Cell Phones": 0.05,
    "Credit Card Payments": 0.05,
    "CarPayments": 0.05,
}

def prompt_user():
    top_priority = input("What is your top priority for the day? ")
    second_priority = input("What is your second priority for the day? ")
    third_priority = input("What is your third priority for the day? ")
    advice = input("Who can you call for advice on achieving your goals? ")
    actions = input("What specific actions do you need to take to move closer to your objectives? ")
    devotional = input("Did you take 10-15 minutes for morning devotional or meditation? (yes/no) ")
    workout = input("Did you complete a 30-45 minute workout or core exercise routine? (yes/no) ")
    budget_review = input("Did you complete your weekly budget review and stick to your budget for the week (assuming an average spending amount of $176.1 per category)? (yes/no) ")
    overspending = {}
    for category in CATEGORIES:
        response = input(f"Did you overspend in {category}? If so, how much? (enter a number or '0' if not applicable) ")
        while not (response.isdigit() or response == "0"):
            response = input(f"Invalid input. Did you overspend in {category}? If so, how much? (enter a number or '0' if not applicable) ")
        overspending[category] = float(response)
    unexpected_expenses = {}
    for category in CATEGORIES:
        response = input(f"Were there any unexpected expenses that came up in {category}? If so, how much? (enter a number or '0' if not applicable) ")
        while not (response.isdigit() or response == "0"):
            response = input(f"Invalid input. Were there any unexpected expenses that came up in {category}? If so, how much? (enter a number or '0' if not applicable) ")
        unexpected_expenses[category] = float(response)
    
    # Calculate the remaining monthly budget for each category based on income and percentages
    remaining_monthly_budget = {}
    total_income = 1174  # Assuming $1174 per week
    for category, percentage in CATEGORIES.items():
        remaining_monthly_budget[category] = round(total_income * percentage * 4, 2)
        if category in overspending:
            remaining_monthly_budget[category] -= overspending[category]
        if category in unexpected_expenses:
            remaining_monthly_budget[category] -= unexpected_expenses[category]

    return (top_priority, second_priority, third_priority, advice, actions, devotional, workout, budget_review, overspending, unexpected_expenses, remaining_monthly_budget)

def write_to_excel(file_path, data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except InvalidFileException:
        workbook = openpyxl.Workbook()
    
    sheet = workbook.active
    row = sheet.max_row + 1
    
    sheet.cell(row=row, column=1, value=data[0])
    sheet.cell(row=row, column=2, value=data[1])
    sheet.cell(row=row, column=3, value=data[2])
    sheet.cell(row=row, column=4, value=data[3])
    sheet.cell(row=row, column=5, value=data[4])
    sheet.cell(row=row, column=6, value=data[5])
    sheet.cell(row=row, column=7, value=data[6])
    sheet.cell(row=row, column=8, value=data[7])
    
    for i, category in enumerate(CATEGORIES):
        sheet.cell(row=row, column=9+i, value=data[8][category])
        sheet.cell(row=row, column=35+i, value=data[9][category])
        sheet.cell(row=row, column=61+i, value=data[10][category])
    
    workbook.save(file_path)
    print("Data written to Excel file successfully!")

# Example usage
data = prompt_user()
write_to_excel("daily_report.xlsx", data)