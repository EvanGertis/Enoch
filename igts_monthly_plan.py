import openpyxl

CATEGORIES = ['Financial goals', 'Tools', 'Deadlines', 
              'Plan', 'Measuring progress']

def prompt_user():
    answers = []
    for category in CATEGORIES:
        if category == 'Deadlines':
            answer = input(f"3. {category}: ")
            while not answer.isnumeric():
                answer = input(f"Please enter a numeric deadline: ")
            answers.append(int(answer))
        elif category == 'Financial goals':
            goal1 = input("Enter your first financial goal for this month: ")
            goal2 = input("Enter your second financial goal for this month: ")
            goal3 = input("Enter your third financial goal for this month: ")
            answers.append(f"{goal1}, {goal2}, {goal3}")
        elif category == 'Measuring progress':
            progress = input("Enter your progress for this month (out of 100): ")
            while not progress.isnumeric() or int(progress) < 0 or int(progress) > 100:
                progress = input("Please enter a valid progress percentage (0-100): ")
            answers.append(int(progress))
        else:      
            answer = input(f"{category}: ")
            answers.append(answer)
    
    # Collect friend's name and meal times
    friend_name = input("Enter your friend's name: ")
    answers.append(friend_name)
    for category in ['Coffee', 'Breakfast', 'Lunch', 'Afternoon Snack', 'Dinner']:
        times = []
        for i in range(1, 6):
            time = input(f"Enter the best time for {category} for person {i}: ")
            times.append(time)
        answers.append(times)
    
    # Collect schedule
    schedule = {}
    for i in range(1, 32):
        date = f"{i}/07/2023" # assuming July 2023
        schedule[date] = {}
        for j, category in enumerate(['Coffee', 'Breakfast', 'Lunch', 'Afternoon Snack', 'Dinner']):
            times = []
            for k, person in enumerate(['Person 1', 'Person 2', 'Person 3', 'Person 4', 'Person 5']):
                time = input(f"Enter the best time for {category} for {person} on {date}: ")
                times.append(time)
            schedule[date][category] = times
        
    answers.append(schedule)
        
    return answers
              
def save_to_excel(answers):
    workbook = openpyxl.Workbook()
    
    # Save financial goals and plan
    sheet1 = workbook.active
    sheet1.title = "Goals and Plan"
    sheet1.cell(row=1, column=1, value="Category")
    sheet1.cell(row=1, column=2, value="Answer")
    for i, category in enumerate(CATEGORIES):
        sheet1.cell(row=i+2, column=1, value=category)
        if category == 'Plan':
            sheet1.cell(row=i+2, column=2, value="See Schedule sheet")
        elif category == 'Measuring progress':
            sheet1.cell(row=i+2, column=2, value=f"{answers[i]}%")
        else:
            sheet1.cell(row=i+2, column=2, value=answers[i])
    
    # Save progress
    progress_sheet = workbook.create_sheet("Progress")
    progress_sheet.cell(row=1, column=1, value="Month")
    progress_sheet.cell(row=1, column=2, value="Progress (out of 100)")
    progress_sheet.cell(row=2, column=1, value="July 2023") # assuming July 2023
    progress_sheet.cell(row=2, column=2, value=f"{answers[-2]}%")
    
    # Save meal plan
    schedule_sheet = workbook.create_sheet("Schedule")
    schedule_sheet.cell(row=1, column=1, value="Date")
    schedule_sheet.cell(row=1, column=2, value="Meal")
    schedule_sheet.cell(row=1, column=3, value="Time")
    schedule_sheet.cell(row=1, column=4, value="Person")
    row = 2
    for date, meals in answers[-1].items():
        for meal, times in meals.items():
            for i, time in enumerate(times):
                schedule_sheet.cell(row=row, column=1, value=date)
                schedule_sheet.cell(row=row, column=2, value=meal)
                schedule_sheet.cell(row=row, column=3, value=time)
                schedule_sheet.cell(row=row, column=4, value=f"Person {i+1}")
                row += 1
    
    workbook.save('monthly_plan.xlsx')

answers = prompt_user()
save_to_excel(answers)