import openpyxl

CATEGORIES = {
    "Emergency Fund": 0.1,
    "Brokerage Account": 0.05,
    "Vacations": 0.05,
    "Entertainment": 0.08,
    "Restaurants": 0.05,
    "Clothing": 0.038,
    "Family": 0.05,
    "Christmas": 0.02,
    "Gifts": 0.02,
    "Hobbies": 0.1,
    "House Upgrades": 0.05,
    "Cable": 0.02,
    "Internet": 0.02,
    "Subscriptions": 0.02,
    "Lawn": 0.02,
    "Beauty Products": 0.02,
    "Gym Memberships": 0.02,
    "Mortgage": 0.18,
    "House Taxes": 0.05,
    "House Insurance": 0.02,
    "Electricity": 0.18,
    "House gas": 0.02,
    "Water / Garbage": 0.02,
    "House Repairs": 0.05,
    "Other housing costs": 0.05,
    "Groceries": 0.1,
    "Vehicle Gas": 0.02,
    "Grandkids": 0.02,
    "Vehicle Taxes and Insurance": 0.05,
    "Vehicle Replacement": 0.05,
    "Oil Changes": 0.02,
    "AAA": 0.02,
    "Tires": 0.02,
    "Vehicle Repairs / Upkeep": 0.05,
    "Health Insurance Premiums": 0.1,
    "Healthcare Costs": 0.05,
    "Cell Phones": 0.05,
    "Credit Card Payments": 0.05,
    "CarPayments": 0.05,
}

def prompt_user():
    top_priority = input("What is your top priority for the day? ")
    second_priority = input("What is your second priority for the day? ")
    third_priority = input("What is your third priority for the day? ")
    advice = input("Who can you call for advice on achieving your goals? ")
    actions = input("What specific actions do you need to take to move closer to your objectives? ")
    devotional = input("Did you take 10-15 minutes for morning devotional or meditation? (yes/no) ")
    workout = input("Did you complete a 30-45 minute workout or core exercise routine? (yes/no) ")
    budget_review = input("Did you complete your weekly budget review and stick to your budget for the week (assuming an average spending amount of $176.1 per category)? (yes/no) ")
    overspending = {}
    for category in CATEGORIES:
        response = input(f"Did you overspend in {category}? If so, how much? (enter a number or '0' if not applicable) ")
        while not (response.isdigit() or response == "0"):
            response = input(f"Invalid input. Did you overspend in {category}? If so, how much? (enter a number or '0' if not applicable) ")
        overspending[category] = float(response)
    unexpected_expenses = {}
    for category in CATEGORIES:
        response = input(f"Were there any unexpected expenses that came up in {category}? If so, how much? (enter a number or '0' if not applicable) ")
        while not (response.isdigit() or response == "0"):
            response = input(f"Invalid input. Were there any unexpected expenses that came up in {category}? If so, how much? (enter a number or '0' if not applicable) ")
        unexpected_expenses[category] = float(response)
    
    # Calculate the remaining monthly budget for each category based on income and percentages
    remaining_monthly_budget = {}
    total_income = 1174  # Assuming $1174 per week
    for category, percentage in CATEGORIES.items():
        remaining_monthly_budget[category] = round(total_income * percentage * 4, 2)
        if category in overspending:
            remaining_monthly_budget[category] -= overspending[category]
        if category in unexpected_expenses:
            remaining_monthly_budget[category] -= unexpected_expenses[category]

    return (top_priority, second_priority, third_priority, advice, actions, devotional, workout, budget_review, overspending, unexpected_expenses, remaining_monthly_budget)

def write_to_excel():
    # Prompt the user for the day's data
    (top_priority, second_priority, third_priority, advice, actions, devotional, workout, budget_review, overspending, unexpected_expenses, remaining_monthly_budget) = prompt_user()

    # Load the workbook and select the correct worksheet
    wb = openpyxl.load_workbook("daily_planner.xlsx")
    ws = wb.active

    # Write the data to the worksheet
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=top_priority)
    ws.cell(row=row, column=2, value=second_priority)
    ws.cell(row=row, column=3, value=third_priority)
    ws.cell(row=row, column=4, value=advice)
    ws.cell(row=row, column=5, value=actions)
    ws.cell(row=row, column=6, value=devotional)
    ws.cell(row=row, column=7, value=workout)
    ws.cell(row=row, column=8, value=budget_review)
    for i, category in enumerate(CATEGORIES.keys(), start=9):
        ws.cell(row=row, column=i, value=remaining_monthly_budget[category])
        ws.cell(row=1, column=i, value=category)
        if category in overspending:
            ws.cell(row=row, column=i+len(CATEGORIES), value=overspending[category])
            ws.cell(row=1, column=i+len(CATEGORIES), value=f"{category} overspending")
        if category in unexpected_expenses:
            ws.cell(row=row, column=i+len(CATEGORIES)*2, value=unexpected_expenses[category])
            ws.cell(row=1, column=i+len(CATEGORIES)*2, value=f"{category} unexpected expenses")

    # Save the workbook
    wb.save("daily_planner.xlsx")

if __name__ == "__main__":
    write_to_excel()