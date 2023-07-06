import openpyxl
from openpyxl import Workbook

# Get the coffee and meal times for 5 different people for each meal
coffee_times = []
breakfast_times = []
lunch_times = []
afternoon_snack_times = []
dinner_times = []

for i in range(5):
    coffee_time = input(f"Enter the best time for coffee for person {i+1}: ")
    coffee_times.append(coffee_time)
    breakfast_time = input(f"Enter the best time for breakfast for person {i+1}: ")
    breakfast_times.append(breakfast_time)
    lunch_time = input(f"Enter the best time for lunch for person {i+1}: ")
    lunch_times.append(lunch_time)
    afternoon_snack_time = input(f"Enter the best time for afternoon snack for person {i+1}: ")
    afternoon_snack_times.append(afternoon_snack_time)
    dinner_time = input(f"Enter the best time for dinner for person {i+1}: ")
    dinner_times.append(dinner_time)

# Create a new Excel workbook
wb = Workbook()

# Create sheets for each section
coffee_sheet = wb.active
coffee_sheet.title = "Coffee"
breakfast_sheet = wb.create_sheet("Breakfast")
lunch_sheet = wb.create_sheet("Lunch")
afternoon_snack_sheet = wb.create_sheet("Afternoon Snack")
dinner_sheet = wb.create_sheet("Dinner")

sheets = [coffee_sheet, breakfast_sheet, lunch_sheet, afternoon_snack_sheet, dinner_sheet]

# Save schedules in each sheet
meal_sections = [
    {"title": "Coffee", "times": coffee_times},
    {"title": "Breakfast", "times": breakfast_times},
    {"title": "Lunch", "times": lunch_times},
    {"title": "Afternoon Snack", "times": afternoon_snack_times},
    {"title": "Dinner", "times": dinner_times},
]

for sheet, section in zip(sheets, meal_sections):
    sheet.cell(row=1, column=1, value=section["title"])
    sheet.cell(row=2, column=1, value="Day")
    sheet.cell(row=2, column=2, value="Time")
    sheet.cell(row=2, column=3, value="Person")

    for day in ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]:
        for i, time in enumerate(section["times"]):
            row = i + 3
            sheet.cell(row=row, column=1, value=day)
            sheet.cell(row=row, column=2, value=time)
            sheet.cell(row=row, column=3, value=f"Person {i+1}")

# Save the workbook to a file
wb.save("weekly_schedule.xlsx")