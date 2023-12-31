import openpyxl
from openpyxl import Workbook
import pywhatkit

# Get the user's location and daily routine
location = input("Enter your location: ")
routine = input("Enter your daily routine (e.g. wake up at 6am, work from 9am to 5pm): ")

# Use pywhatkit to search for optimal meal times
coffee_time = pywhatkit.search(f"best time for coffee in {location} {routine}")
breakfast_time = pywhatkit.search(f"best time for breakfast in {location} {routine}")
lunch_time = pywhatkit.search(f"best time for lunch in {location} {routine}")
snack_time = pywhatkit.search(f"best time for snack in {location} {routine}")
dinner_time = pywhatkit.search(f"best time for dinner in {location} {routine}")

# Create a new Excel workbook
wb = Workbook()

# Create sheets for each section
contacts_sheet = wb.active
contacts_sheet.title = "Contacts"

numerical_results_sheet = wb.create_sheet("Numerical Results")

# Save contacts in the contacts sheet
contact_sections = [
    {"title": "Coffee", "time": coffee_time, "start_row": 2},
    {"title": "Breakfast", "time": breakfast_time, "start_row": 22},
    {"title": "Lunch", "time": lunch_time, "start_row": 42},
    {"title": "Snack", "time": snack_time, "start_row": 62},
    {"title": "Dinner", "time": dinner_time, "start_row": 82},
]

for section in contact_sections:
    start_row = section["start_row"]
    contacts_sheet.cell(row=start_row, column=1, value=section["title"])
    contacts_sheet.cell(row=start_row, column=2, value="Week")
    contacts_sheet.cell(row=start_row, column=3, value="Day")
    contacts_sheet.cell(row=start_row, column=4, value="Time")
    contacts_sheet.cell(row=start_row, column=5, value="Name")

    for week in range(1, 5):
        for day in ["MON", "TUE", "WED", "THUR", "FRI"]:
            start_row += 1
            contacts_sheet.cell(row=start_row, column=2, value=week)
            contacts_sheet.cell(row=start_row, column=3, value=day)
            contacts_sheet.cell(row=start_row, column=4, value=section["time"])

# Save numerical results in the numerical results sheet
numerical_results_sheet.cell(row=1, column=1, value="Progress Check")
numerical_results_sheet.cell(row=1, column=2, value="Tool Name")

tools = [
    "RYA budget sheet",
    "Wave",
    "Mint",
    "QuickBooks",
    "Vanguard",
    "Robinhood",
    "TD Ameritrade",
    "Schwab",
    "BlockFi",
]

for index, tool in enumerate(tools, start=2):
    numerical_results_sheet.cell(row=index, column=2, value=tool)

# Save the workbook to a file
wb.save("monthly_plan.xlsx")