import openpyxl
from openpyxl import Workbook

# Get the user's and friend's coffee and meal times
user_coffee_time = input("Enter the best time for coffee for you: ")
user_meal_time = input("Enter the best time for a meal for you: ")
friend_coffee_time = input("Enter the best time for coffee for your friend: ")
friend_meal_time = input("Enter the best time for a meal for your friend: ")
friend_name = input("Enter your friend's name: ")

# Create a new Excel workbook
wb = Workbook()

# Create sheets for each section
contacts_sheet = wb.active
contacts_sheet.title = "Contacts"

numerical_results_sheet = wb.create_sheet("Numerical Results")

# Save contacts in the contacts sheet
contact_sections = [
    {"title": "Coffee", "time": user_coffee_time, "friend_time": friend_coffee_time, "start_row": 2},
    {"title": "Meal", "time": user_meal_time, "friend_time": friend_meal_time, "start_row": 22},
]

for section in contact_sections:
    start_row = section["start_row"]
    contacts_sheet.cell(row=start_row, column=1, value=section["title"])
    contacts_sheet.cell(row=start_row, column=2, value="Week")
    contacts_sheet.cell(row=start_row, column=3, value="Day")
    contacts_sheet.cell(row=start_row, column=4, value="Time")
    contacts_sheet.cell(row=start_row, column=5, value="Name")

    for week in range(1, 5):
        for day in ["MON", "TUE", "WED", "THU", "FRI"]:
            start_row += 1
            contacts_sheet.cell(row=start_row, column=2, value=week)
            contacts_sheet.cell(row=start_row, column=3, value=day)
            contacts_sheet.cell(row=start_row, column=4, value=section["time"] + " with " + friend_name + " at " + section["friend_time"])

# Save numerical results in the numerical results sheet
numerical_results_sheet.cell(row=1, column=1, value="Progress Check")
numerical_results_sheet.cell(row=1, column=2, value="Tool Name")

tools = [
    "RYA budget sheet",
    "Wave",
    "Mint",
    "QuickBooks",
    "Vanguard",
    "Robinhood",
    "TD Ameritrade",
    "Schwab",
    "BlockFi",
]

for index, tool in enumerate(tools, start=2):
    numerical_results_sheet.cell(row=index, column=2, value=tool)

# Save the workbook to a file
wb.save("monthly_plan.xlsx")