import openpyxl

def prompt_user():
    top_priority = input("What is your top priority for the day? ")
    financial_data = financial_questions()
    advice = input("Who can you call for advice on achieving your goals? ")
    actions = input("What specific actions do you need to take to move closer to your objectives? ")
    devotional = input("Did you take 10-15 minutes for morning devotional or meditation? ")
    workout = input("Did you complete a 30-45 minute workout or core exercise routine? ")
    budget_review = input("Did you complete your weekly budget review and stick to your budget for the week (assuming an average spending amount of $176.1 per category)? ")
    overspending = {}
    for category in ["House Taxes", "House Insurance", "Electricity", "House Gas", "Water/Garbage", "House Repairs", "Other Housing Costs", "Groceries", "Vehicle Gas", "Grandkids", "Vehicle Taxes and Insurance", "Vehicle Replacement", "Oil Changes", "AAA", "Tires", "Vehicle Repairs/Upkeep", "Health Insurance Premiums", "Healthcare Costs", "Cell Phones", "Mortgage", "Car Payments", "Other Debt Payments", "Giving", "Emergency Fund", "Brokerage Account", "Vacations", "Entertainment", "Restaurants", "Clothing", "Family", "Christmas", "Gifts", "Hobbies", "House Upgrades", "Cable", "Internet", "Subscriptions", "Lawn", "Beauty Products", "Gym Memberships", "Pets", "Life Insurance", "Disability Income Insurance", "Tax-Free Savings Bucket", "Medical Expenses/Hospital", "Lawyers", "Other"]:
        overspending[category] = input(f"Did you overspend in {category}? If so, how much?")
    unexpected_expenses = {}
    for category in ["House Taxes", "House Insurance", "Electricity", "House Gas", "Water/Garbage", "House Repairs", "Other Housing Costs", "Groceries", "Vehicle Gas", "Grandkids", "Vehicle Taxes and Insurance", "Vehicle Replacement", "Oil Changes", "AAA", "Tires", "Vehicle Repairs/Upkeep", "Health Insurance Premiums", "Healthcare Costs", "Cell Phones", "Mortgage", "Car Payments", "Other Debt Payments", "Giving", "Emergency Fund", "Brokerage Account", "Vacations", "Entertainment", "Restaurants", "Clothing", "Family", "Christmas", "Gifts", "Hobbies", "House Upgrades", "Cable", "Internet", "Subscriptions", "Lawn", "Beauty Products", "Gym Memberships", "Pets", "Life Insurance", "Disability Income Insurance", "Tax-Free Savings Bucket", "Medical Expenses/Hospital", "Lawyers", "Other"]:
        unexpected_expenses[category] = input(f"Were there any unexpected expenses that came up in {category}? If so, how much?")
    remaining_monthly_budget = {}
    for category in ["House Taxes", "House Insurance", "Electricity", "House Gas", "Water/Garbage", "House Repairs", "Other Housing Costs", "Groceries", "Vehicle Gas", "Grandkids", "Vehicle Taxes and Insurance", "Vehicle Replacement", "Oil Changes", "AAA", "Tires", "Vehicle Repairs/Upkeep", "Health Insurance Premiums", "Healthcare Costs", "Cell Phones", "Mortgage", "Car Payments", "Other Debt Payments", "Giving", "Emergency Fund", "Brokerage Account", "Vacations", "Entertainment", "Restaurants", "Clothing", "Family", "Christmas", "Gifts", "Hobbies", "House Upgrades", "Cable", "Internet", "Subscriptions", "Lawn", "Beauty Products", "Gym Memberships", "Pets", "Life Insurance", "Disability Income Insurance", "Tax-Free Savings Bucket", "Medical Expenses/Hospital", "Lawyers", "Other"]:
        remaining_monthly_budget[category] = input(f"How much money do you have left for the rest of the month in {category}?")

    return top_priority, financial_data, advice, actions, devotional, workout, budget_review, overspending, unexpected_expenses, remaining_monthly_budget

def financial_questions():
    money_made_today = input("How much moneydid you make today? ")
    money_spent_today = input("How much money did you spend today? ")
    money_saved_today = input("How much money did you save today? ")
    money_invested_today = input("How much money did you invest today? ")
    return money_made_today, money_spent_today, money_saved_today, money_invested_today

def create_excel_worksheet():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Daily Summary"
    worksheet["A1"] = "Top Priority"
    worksheet["B1"] = "Money Made Today"
    worksheet["C1"] = "Money Spent Today"
    worksheet["D1"] = "Money Saved Today"
    worksheet["E1"] = "Money Invested Today"
    worksheet["F1"] = "Advice Contact"
    worksheet["G1"] = "Actions Needed"
    worksheet["H1"] = "Morning Devotional"
    worksheet["I1"] = "Workout Completed"
    worksheet["J1"] = "Weekly Budget Review"
    worksheet["K1"] = "Overspending"
    worksheet["L1"] = "Unexpected Expenses"
    worksheet["M1"] = "Remaining Monthly Budget"
    row_num = 2
    while True:
        top_priority, financial_data, advice, actions, devotional, workout, budget_review, overspending, unexpected_expenses, remaining_monthly_budget = prompt_user()
        if top_priority == "":
            break
        worksheet.cell(row=row_num, column=1).value = top_priority
        worksheet.cell(row=row_num, column=2).value = financial_data[0]
        worksheet.cell(row=row_num, column=3).value = financial_data[1]
        worksheet.cell(row=row_num, column=4).value = financial_data[2]
        worksheet.cell(row=row_num, column=5).value = financial_data[3]
        worksheet.cell(row=row_num, column=6).value = advice
        worksheet.cell(row=row_num, column=7).value = actions
        worksheet.cell(row=row_num, column=8).value = devotional
        worksheet.cell(row=row_num, column=9).value = workout
        worksheet.cell(row=row_num, column=10).value = budget_review
        for i, category in enumerate(["House Taxes", "House Insurance", "Electricity", "House Gas", "Water/Garbage", "House Repairs", "Other Housing Costs", "Groceries", "Vehicle Gas", "Grandkids", "Vehicle Taxes and Insurance", "Vehicle Replacement", "Oil Changes", "AAA", "Tires", "Vehicle Repairs/Upkeep", "Health Insurance Premiums", "Healthcare Costs", "Cell Phones", "Mortgage", "Car Payments", "Other Debt Payments", "Giving", "Emergency Fund", "Brokerage Account", "Vacations", "Entertainment", "Restaurants", "Clothing", "Family", "Christmas", "Gifts", "Hobbies", "House Upgrades", "Cable", "Internet", "Subscriptions", "Lawn", "Beauty Products", "Gym Memberships", "Pets", "Life Insurance", "Disability Income Insurance", "Tax-Free Savings Bucket", "Medical Expenses/Hospital", "Lawyers", "Other"]):
            worksheet.cell(row=row_num, column=11+i).value = overspending[category]
            worksheet.cell(row=row_num, column=61+i).value = unexpected_expenses[category]
            worksheet.cell(row=row_num, column=111+i).value = remaining_monthly_budget[category]
        row_num += 1
    workbook.save("daily_summary.xlsx")

create_excel_worksheet()